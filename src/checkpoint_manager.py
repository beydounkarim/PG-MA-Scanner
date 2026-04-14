"""
Progressive checkpoint manager for scan resume safety.

Saves scan state after each major step so the pipeline can resume
from where it left off if the process is interrupted.
Also provides Excel backup output.
"""

import json
import os
import sys
from datetime import date, datetime, timedelta
from typing import Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# Add src to path for sibling imports
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from sheets_output import get_deal_date, get_date_group, get_date_group_order, get_group_header


CHECKPOINT_DIR = "data/checkpoints"


def make_checkpoint_path(label: str = "") -> str:
    """Generate a timestamped checkpoint file path."""
    os.makedirs(CHECKPOINT_DIR, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    suffix = f"_{label}" if label else ""
    return os.path.join(CHECKPOINT_DIR, f"scan{suffix}_{ts}.json")


def save_progressive(path: str, data: dict) -> None:
    """Atomically write checkpoint data to disk."""
    os.makedirs(os.path.dirname(path), exist_ok=True)
    tmp = path + ".tmp"
    with open(tmp, "w") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)
    os.replace(tmp, path)


def load_checkpoint(path: str) -> dict:
    """Load a previously saved checkpoint."""
    with open(path, "r") as f:
        return json.load(f)


def mark_step(ckpt: dict, step: str) -> None:
    """Mark a pipeline step as completed (in-memory, call save_progressive to persist)."""
    if "completed_steps" not in ckpt:
        ckpt["completed_steps"] = []
    if step not in ckpt["completed_steps"]:
        ckpt["completed_steps"].append(step)


def is_step_done(ckpt: dict, step: str) -> bool:
    """Check whether a pipeline step was already completed."""
    return step in ckpt.get("completed_steps", [])


# ---------------------------------------------------------------------------
# Excel backup
# ---------------------------------------------------------------------------

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
SECTION_FONT = Font(bold=True, color="333333", size=11)
SECTION_FILL = PatternFill(start_color="E5E5E5", end_color="E5E5E5", fill_type="solid")
WRAP = Alignment(wrap_text=True, vertical="top")


def _write_header(ws, headers):
    ws.append(headers)
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = WRAP


def _write_section_header(ws, text, num_cols):
    """Write a date group section header row in Excel."""
    ws.append([text] + [""] * (num_cols - 1))
    row_num = ws.max_row
    for col_idx in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=col_idx)
        cell.font = SECTION_FONT
        cell.fill = SECTION_FILL
    ws.merge_cells(start_row=row_num, start_column=1,
                   end_row=row_num, end_column=num_cols)


def _write_grouped_deals(ws, deals, headers, row_builder, today=None):
    """Write deals grouped by date with section headers."""
    if today is None:
        today = date.today()

    groups = get_date_group_order(today)
    num_cols = len(headers)

    # Group deals
    grouped = {}
    for d in deals:
        deal_d = get_deal_date(d)
        group = get_date_group(deal_d, today)
        grouped.setdefault(group, []).append(d)

    # Sort each group by date descending
    for group in grouped:
        grouped[group].sort(
            key=lambda d: get_deal_date(d) or date.min, reverse=True
        )

    # Write each group
    for group in groups:
        header_text = get_group_header(group, today)
        _write_section_header(ws, header_text, num_cols)
        for d in grouped.get(group, []):
            ws.append(row_builder(d))


def save_excel_backup(new_deals: list[dict],
                      excluded_deals: list[dict],
                      unverified_deals: list[dict],
                      filepath: str) -> str:
    """
    Save all scan results to an Excel workbook with three sheets.
    Deals are grouped by date with section headers.

    Returns the filepath written.
    """
    today = date.today()
    wb = openpyxl.Workbook()

    # --- New Deals sheet ---
    deal_headers = [
        "Acquiror", "Target", "Deal Type", "Deal Status", "Description",
        "Deal Value", "Date Rumored", "Date Announced", "Date Closed",
        "Sector", "Source", "Source Link", "PG Account", "Clean Name",
        "Opportunity", "Source Validation",
    ]
    ws = wb.active
    ws.title = "New Deals"
    _write_header(ws, deal_headers)

    def deal_row(d):
        return [
            d.get("acquiror", ""),
            d.get("target", ""),
            d.get("deal_type", ""),
            d.get("deal_status", ""),
            d.get("description", ""),
            d.get("deal_value", ""),
            d.get("date_rumor", ""),
            d.get("date_announced", ""),
            d.get("date_closed", ""),
            d.get("sector", ""),
            d.get("source", ""),
            d.get("source_link", ""),
            d.get("pg_account_name", ""),
            d.get("clean_name", ""),
            d.get("opportunity", ""),
            d.get("source_validation", ""),
        ]

    _write_grouped_deals(ws, new_deals, deal_headers, deal_row, today)

    # --- Excluded sheet ---
    exc_headers = [
        "Acquiror", "Target", "Deal Value", "Sector",
        "Description", "Exclusion Reason",
    ]
    ws_exc = wb.create_sheet("Excluded")
    _write_header(ws_exc, exc_headers)
    for d in excluded_deals:
        ws_exc.append([
            d.get("acquiror", ""),
            d.get("target", ""),
            d.get("deal_value", ""),
            d.get("sector", ""),
            d.get("description", ""),
            d.get("exclusion_reason", ""),
        ])

    # --- Unverified sheet ---
    unv_headers = [
        "Acquiror", "Target", "Deal Status", "Sector",
        "Description", "Source Link", "Failure Reason",
    ]
    ws_unv = wb.create_sheet("Unverified")
    _write_header(ws_unv, unv_headers)
    for d in unverified_deals:
        ws_unv.append([
            d.get("acquiror", ""),
            d.get("target", ""),
            d.get("deal_status", ""),
            d.get("sector", ""),
            d.get("description", ""),
            d.get("source_link", ""),
            d.get("validation_failure_reason", ""),
        ])

    # Auto-width columns (approximate)
    for sheet in wb.sheetnames:
        ws_tmp = wb[sheet]
        for col in ws_tmp.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=10)
            ws_tmp.column_dimensions[col[0].column_letter].width = min(max_len + 2, 60)

    # Freeze header row on all sheets
    for sheet in wb.sheetnames:
        wb[sheet].freeze_panes = "A2"

    os.makedirs(os.path.dirname(filepath) if os.path.dirname(filepath) else ".", exist_ok=True)
    wb.save(filepath)
    return filepath
