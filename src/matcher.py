"""
Company list loader and fuzzy matching module.

Loads PG customer accounts from Excel and provides fuzzy matching
to map deal company names to PG accounts.
"""

import openpyxl
from thefuzz import fuzz
from typing import Optional


def load_company_list(filepath: str) -> list[dict]:
    """
    Load PG_Acct_List.xlsx and return deduplicated list of companies.

    Args:
        filepath: Path to PG_Acct_List.xlsx

    Returns:
        List of dicts with keys: account_name, clean_name
        Example: [{"account_name": "Chevron | NA", "clean_name": "Chevron"}, ...]

    Raises:
        FileNotFoundError: If Excel file doesn't exist
        ValueError: If required sheets or columns are missing
    """
    try:
        workbook = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    except FileNotFoundError:
        raise FileNotFoundError(
            f"Company list not found at {filepath}. "
            f"Please ensure PG_Acct_List.xlsx is in the data/ directory."
        )

    companies = []
    seen_clean_names = set()

    # Load from both A&B Customers and C&D Customers sheets
    sheet_tier_map = {'A&B Customers': 'AB', 'C&D Customers': 'CD'}
    for sheet_name, tier in sheet_tier_map.items():
        if sheet_name not in workbook.sheetnames:
            raise ValueError(
                f"Required sheet '{sheet_name}' not found in {filepath}. "
                f"Available sheets: {workbook.sheetnames}"
            )

        sheet = workbook[sheet_name]

        # Find header row and column indices
        header_row = None
        account_name_col = None
        clean_name_col = None

        for row_idx, row in enumerate(sheet.iter_rows(max_row=10, values_only=True), start=1):
            if row and any(cell for cell in row):
                # Look for "Account Name" and "Clean Name" columns
                for col_idx, cell in enumerate(row):
                    if cell and isinstance(cell, str):
                        if 'account' in cell.lower() and 'name' in cell.lower():
                            account_name_col = col_idx
                        elif 'clean' in cell.lower() and 'name' in cell.lower():
                            clean_name_col = col_idx

                if account_name_col is not None and clean_name_col is not None:
                    header_row = row_idx
                    break

        if header_row is None:
            raise ValueError(
                f"Could not find 'Account Name' and 'Clean Name' columns in sheet '{sheet_name}'"
            )

        # Read data rows
        for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
            if not row or not any(row):
                continue

            account_name = row[account_name_col] if account_name_col < len(row) else None
            clean_name = row[clean_name_col] if clean_name_col < len(row) else None

            if account_name and clean_name:
                account_name = str(account_name).strip()
                clean_name = str(clean_name).strip()

                # Deduplicate on clean_name (keep first occurrence)
                if clean_name and clean_name not in seen_clean_names:
                    seen_clean_names.add(clean_name)
                    companies.append({
                        "account_name": account_name,
                        "clean_name": clean_name,
                        "tier": tier
                    })

    workbook.close()

    if not companies:
        raise ValueError(f"No companies loaded from {filepath}")

    return companies


def split_by_tier(companies: list[dict]) -> tuple[list[dict], list[dict]]:
    """
    Split companies into A&B and C&D tiers.

    Args:
        companies: List of company dicts with 'tier' key

    Returns:
        Tuple of (ab_companies, cd_companies)
    """
    ab = [c for c in companies if c.get("tier") == "AB"]
    cd = [c for c in companies if c.get("tier") == "CD"]
    return ab, cd


def fuzzy_match(company_name: str, company_list: list[dict], threshold: int = 75) -> Optional[dict]:
    """
    Find the best fuzzy match for a company name in the PG company list.

    Args:
        company_name: Company name from deal data (e.g., "Chevron Corporation")
        company_list: List of dicts from load_company_list()
        threshold: Minimum similarity score (0-100). Default: 75

    Returns:
        Matched dict with account_name and clean_name, or None if no match above threshold

    Examples:
        >>> companies = [{"account_name": "Chevron | NA", "clean_name": "Chevron"}]
        >>> fuzzy_match("Chevron Corporation", companies)
        {"account_name": "Chevron | NA", "clean_name": "Chevron"}
        >>> fuzzy_match("RandomCompany", companies)
        None
    """
    if not company_name or not company_list:
        return None

    company_name = str(company_name).strip()
    if not company_name:
        return None

    # Normalize the input name (strip common suffixes)
    normalized_input = normalize_company_name(company_name)

    best_match = None
    best_score = 0

    for company in company_list:
        clean_name = company.get("clean_name", "")
        if not clean_name:
            continue

        # Try matching with normalized names first (helps with "Rio Tinto Limited" → "Rio Tinto")
        normalized_clean = normalize_company_name(clean_name)
        normalized_score = fuzz.token_sort_ratio(normalized_input, normalized_clean)

        # Also try original names (in case normalization hurts the match)
        original_score = fuzz.token_sort_ratio(company_name.lower(), clean_name.lower())

        # Try partial matching - if input contains the clean name, boost score
        # (helps with "INEOS Energy" → "INEOS", "Glencore plc" → "Glencore")
        partial_score = 0
        if clean_name.lower() in company_name.lower():
            # Full clean name appears in input - very likely a match
            partial_score = 90
        elif normalized_clean and normalized_clean in normalized_input:
            # Normalized clean name appears in normalized input
            partial_score = 85

        # Use the best score from all methods
        score = max(normalized_score, original_score, partial_score)

        if score > best_score:
            best_score = score
            best_match = company

    # Return match only if above threshold
    if best_score >= threshold:
        return best_match

    return None


def normalize_company_name(name: str) -> str:
    """
    Normalize a company name for comparison.

    Strips common suffixes like Corp, Corporation, Inc, Ltd, LLC.
    Converts to lowercase and removes extra whitespace.

    Args:
        name: Company name to normalize

    Returns:
        Normalized company name

    Examples:
        >>> normalize_company_name("Chevron Corporation")
        "chevron"
        >>> normalize_company_name("BHP Group Ltd.")
        "bhp group"
    """
    if not name:
        return ""

    name = str(name).strip().lower()

    # Remove common suffixes
    suffixes = [
        " corporation", " corp.", " corp",
        " incorporated", " inc.", " inc",
        " limited", " ltd.", " ltd",
        " llc", " l.l.c.", " l.l.c",
        " plc", " p.l.c.",
        " sa", " s.a.",
        " se", " s.e.",
        " nv", " n.v.",
        " ag", " a.g.",
        " gmbh",
        " holdings",
        " group"
    ]

    for suffix in suffixes:
        if name.endswith(suffix):
            name = name[:-len(suffix)].strip()
            break  # Only remove one suffix

    # Remove extra whitespace
    name = ' '.join(name.split())

    return name
