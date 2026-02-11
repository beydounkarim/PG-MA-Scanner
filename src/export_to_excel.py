#!/usr/bin/env python3
"""
Export to Excel - Create an Excel version of the Google Sheet with formatting.

Usage:
    python src/export_to_excel.py
    python src/export_to_excel.py --output data/deals.xlsx
"""

import argparse
import os
import sys
from datetime import datetime

# Add src to path
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from dotenv import load_dotenv
load_dotenv()

from sheets_output import open_sheet

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Error: openpyxl is required for Excel export")
    print("Install it with: pip install openpyxl")
    sys.exit(1)


def export_to_excel(output_path: str = None):
    """
    Export Google Sheet to Excel with matching formatting.

    Args:
        output_path: Path to save Excel file (default: data/PG_MA_Deals.xlsx)
    """
    if output_path is None:
        output_path = os.path.join(
            os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
            "data", "PG_MA_Deals.xlsx"
        )

    # Ensure output directory exists
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    print("Connecting to Google Sheets...")
    spreadsheet = open_sheet()
    ws_source = spreadsheet.worksheet('Deals')

    # Get all data
    all_rows = ws_source.get_all_values()
    print(f"✓ Loaded {len(all_rows)} rows from Google Sheets\n")

    # Create Excel workbook
    print("Creating Excel workbook...")
    wb = Workbook()
    ws = wb.active
    ws.title = "Deals"

    # Define styles
    title_font = Font(name='Calibri', size=14, bold=True, color='000000')
    title_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    title_alignment = Alignment(horizontal='center', vertical='center')

    header_font = Font(name='Calibri', size=11, bold=True, color='000000')
    header_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    header_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    section_font = Font(name='Calibri', size=11, bold=True, color='333333')
    section_fill = PatternFill(start_color='E6E6E6', end_color='E6E6E6', fill_type='solid')

    normal_font = Font(name='Calibri', size=10, color='000000')
    normal_alignment = Alignment(horizontal='left', vertical='top', wrap_text=False)

    # Alternating row fills
    white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    light_gray_fill = PatternFill(start_color='FAFAFA', end_color='FAFAFA', fill_type='solid')

    # Thin border
    thin_border = Border(
        left=Side(style='thin', color='E0E0E0'),
        right=Side(style='thin', color='E0E0E0'),
        top=Side(style='thin', color='E0E0E0'),
        bottom=Side(style='thin', color='E0E0E0')
    )

    print("Writing data to Excel...")

    # Write all data
    for row_idx, row in enumerate(all_rows, start=1):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value

            # Apply formatting based on row type
            if row_idx == 1:
                # Title row
                cell.font = title_font
                cell.fill = title_fill
                cell.alignment = title_alignment
            elif row and row[0] == 'PG Account Name':
                # Header row
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            elif row and row[0].startswith('═══'):
                # Section header
                cell.font = section_font
                cell.fill = section_fill
                cell.alignment = Alignment(horizontal='left', vertical='center')
            elif row and any(row):
                # Data row
                cell.font = normal_font
                cell.alignment = normal_alignment
                cell.border = thin_border

                # Alternating row colors (skip empty rows and section headers)
                if not row[0].startswith('═══') and row[0] != '':
                    if row_idx % 2 == 0:
                        cell.fill = light_gray_fill
                    else:
                        cell.fill = white_fill

        if row_idx % 100 == 0:
            print(f"  Processed {row_idx}/{len(all_rows)} rows...")

    print(f"✓ Wrote {len(all_rows)} rows\n")

    # Set column widths
    print("Adjusting column widths...")
    column_widths = {
        'A': 25,  # PG Account Name
        'B': 20,  # Clean Name
        'C': 30,  # Acquiror
        'D': 30,  # Target
        'E': 15,  # Deal Status
        'F': 20,  # Sector
        'G': 50,  # Description
        'H': 15,  # Date of Rumor
        'I': 18,  # Date of Announcement
        'J': 15,  # Date Closed
        'K': 15,  # Deal Value
        'L': 20,  # Source
        'M': 40,  # Source Link
        'N': 50,  # Potential Opportunity for PG
        'O': 18,  # Source Validation
        'P': 20,  # deal_id
        'Q': 18,  # stages_reported
        'R': 15,  # first_seen
        'S': 15,  # last_updated
        'T': 25,  # scan_period
    }

    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Freeze panes at header row (row 3)
    print("Freezing header row...")
    ws.freeze_panes = 'A4'  # Freeze rows 1-3

    # Auto-filter on header row
    print("Adding auto-filter...")
    # Find header row
    header_row = None
    for row_idx, row in enumerate(all_rows, start=1):
        if row and row[0] == 'PG Account Name':
            header_row = row_idx
            break

    if header_row:
        ws.auto_filter.ref = f'A{header_row}:T{len(all_rows)}'

    # Save workbook
    print(f"\nSaving Excel file to: {output_path}")
    wb.save(output_path)

    # Get file size
    file_size = os.path.getsize(output_path)
    file_size_mb = file_size / (1024 * 1024)

    print()
    print("=" * 70)
    print("EXCEL EXPORT COMPLETE")
    print("=" * 70)
    print(f"File saved:    {output_path}")
    print(f"File size:     {file_size_mb:.2f} MB")
    print(f"Total rows:    {len(all_rows)}")
    print(f"Total columns: {len(all_rows[0]) if all_rows else 0}")
    print()
    print("Features:")
    print("  ✓ Professional formatting matching Google Sheet")
    print("  ✓ Frozen header row for easy scrolling")
    print("  ✓ Auto-filter enabled on all columns")
    print("  ✓ Alternating row colors for readability")
    print("  ✓ Optimized column widths")
    print("  ✓ Section headers formatted")


def main():
    """Main entry point."""
    parser = argparse.ArgumentParser(
        description="Export Google Sheet to Excel with formatting"
    )

    parser.add_argument(
        '--output',
        help='Output Excel file path (default: data/PG_MA_Deals.xlsx)'
    )

    args = parser.parse_args()

    print("=" * 70)
    print("EXPORT GOOGLE SHEET TO EXCEL")
    print("=" * 70)
    print()

    try:
        export_to_excel(output_path=args.output)
        return 0
    except Exception as e:
        print(f"\n✗ Error: {e}", file=sys.stderr)
        import traceback
        traceback.print_exc()
        return 1


if __name__ == "__main__":
    sys.exit(main())
