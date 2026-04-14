"""
Enrich PitchBook Industrial Software Transactions with AI-generated
deal descriptions and OFFENSIVE/DEFENSIVE opportunity classifications.

Phase 1: Fuzzy-match against existing scanner output to copy enrichment.
Phase 2: AI enrichment via Claude API for unmatched deals.
Phase 3: Write enriched Excel output.

Usage:
    venv/bin/python src/enrich_pitchbook.py \
        --input data/Industrial_Software_Transactions.xlsx \
        --scanner data/PG_MA_Scanner_Deals.xlsx \
        --output data/Industrial_Software_Enriched.xlsx \
        [--skip-ai]           # Matching only, no API calls
        [--resume CHECKPOINT]  # Resume from checkpoint
        [--test N]            # Process first N rows only
        [--workers N]         # Parallel workers (default 1)
"""

import argparse
import json
import os
import re
import sys
import time
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, date
from typing import Optional

import anthropic
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from thefuzz import fuzz

# Add src to path for sibling imports
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from scanner import (
    call_claude_with_retry,
    extract_json_from_response,
    extract_text_from_response,
    clean_citation_tags,
    is_fatal_api_error,
)
from matcher import normalize_company_name

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

CHECKPOINT_DIR = "data/checkpoints"

# PitchBook header is row 3 (1-indexed), data starts row 4
PB_HEADER_ROW = 3
PB_DATA_START_ROW = 4

# New columns to append (AC-AH)
NEW_HEADERS = [
    "Description",
    "Opportunity",
    "Verification Link",
    "Source Validation",
    "Enrichment Source",
    "QA Status",
]

# Matching thresholds
TARGET_MATCH_THRESHOLD = 80
ACQUIROR_MATCH_THRESHOLD = 75
HIGH_CONFIDENCE_SCORE = 85
DATE_PROXIMITY_DAYS = 90

# AI settings
AI_MODEL = "claude-sonnet-4-20250514"
AI_SLEEP_BETWEEN_CALLS = 2
CHECKPOINT_EVERY_N = 25


# ---------------------------------------------------------------------------
# Phase 1: Data Loading
# ---------------------------------------------------------------------------

def load_pitchbook(filepath: str) -> tuple[list[list], list[str]]:
    """Load PitchBook Excel file.

    Returns:
        (data_rows, headers) where data_rows is a list of lists (one per data row)
        and headers is the column header list from row 3.
    """
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active

    headers = []
    data_rows = []

    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i == PB_HEADER_ROW:
            headers = [str(c) if c is not None else "" for c in row]
        elif i >= PB_DATA_START_ROW:
            data_rows.append(list(row))

    wb.close()
    print(f"  Loaded {len(data_rows)} PitchBook rows, {len(headers)} columns")
    return data_rows, headers


def load_scanner_deals(filepath: str) -> list[dict]:
    """Load scanner deals from Excel, skipping section header rows (═══)."""
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb["New Deals"]

    headers = []
    deals = []

    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i == 1:
            headers = [str(c) if c is not None else "" for c in row]
            continue

        # Skip section header rows
        first_cell = str(row[0]) if row[0] is not None else ""
        if first_cell.startswith("═══") or first_cell.startswith("==="):
            continue

        # Skip empty rows
        if not any(c for c in row):
            continue

        deal = {}
        for j, header in enumerate(headers):
            val = row[j] if j < len(row) else None
            deal[header] = str(val) if val is not None else ""
        deals.append(deal)

    wb.close()
    print(f"  Loaded {len(deals)} scanner deals")
    return deals


# ---------------------------------------------------------------------------
# Phase 1: Acquiror Extraction
# ---------------------------------------------------------------------------

def extract_acquiror_from_investors(investors_str: str) -> str:
    """Extract the primary acquiror from PitchBook Investors column.

    Handles:
    - Stock tickers: "(STO: XXX)" → removed
    - Person names in parens: "(John Smith)" → removed
    - Comma-separated list → take first entry
    - Strips whitespace
    """
    if not investors_str or investors_str == "None":
        return ""

    text = str(investors_str).strip()

    # Remove stock ticker patterns like (STO: XXX), (NAS: XXX), (NYSE: XXX), etc.
    text = re.sub(r'\([A-Z]{2,5}:\s*[A-Z0-9.]+\)', '', text)

    # Remove parenthetical person names (single words or "First Last" patterns)
    # but keep company names in parens (they usually have Corp/Inc/etc.)
    text = re.sub(r'\([A-Z][a-z]+ [A-Z][a-z]+\)', '', text)

    # Take first comma-separated entry
    parts = [p.strip() for p in text.split(",")]
    first = parts[0].strip() if parts else ""

    # Clean up any remaining artifacts
    first = re.sub(r'\s+', ' ', first).strip()

    return first


# ---------------------------------------------------------------------------
# Phase 1: Fuzzy Matching
# ---------------------------------------------------------------------------

def parse_deal_date(date_val) -> Optional[date]:
    """Parse a date value from either PitchBook or scanner format."""
    if date_val is None:
        return None
    if isinstance(date_val, datetime):
        return date_val.date()
    if isinstance(date_val, date):
        return date_val
    try:
        return date.fromisoformat(str(date_val).strip()[:10])
    except (ValueError, AttributeError):
        return None


def compute_match_score(pb_target: str, pb_acquiror: str, pb_date,
                        scanner_deal: dict) -> float:
    """Compute a confidence score (0-100) for matching a PB row to a scanner deal."""
    scanner_target = scanner_deal.get("Target", "")
    scanner_acquiror = scanner_deal.get("Acquiror", "")

    # Normalize names
    pb_target_norm = normalize_company_name(pb_target)
    scanner_target_norm = normalize_company_name(scanner_target)
    pb_acq_norm = normalize_company_name(pb_acquiror)
    scanner_acq_norm = normalize_company_name(scanner_acquiror)

    # Target similarity
    target_score = fuzz.token_sort_ratio(pb_target_norm, scanner_target_norm) if pb_target_norm and scanner_target_norm else 0

    if target_score < TARGET_MATCH_THRESHOLD:
        return 0.0

    # Acquiror similarity
    acq_score = fuzz.token_sort_ratio(pb_acq_norm, scanner_acq_norm) if pb_acq_norm and scanner_acq_norm else 0

    if acq_score < ACQUIROR_MATCH_THRESHOLD:
        return 0.0

    # Base score from name matching (weighted: target 60%, acquiror 40%)
    base_score = (target_score * 0.6) + (acq_score * 0.4)

    # Date proximity bonus
    pb_d = parse_deal_date(pb_date)
    scanner_d = None
    for date_key in ["Date Closed", "Date Announced", "Date Rumored"]:
        scanner_d = parse_deal_date(scanner_deal.get(date_key))
        if scanner_d:
            break

    if pb_d and scanner_d:
        days_apart = abs((pb_d - scanner_d).days)
        if days_apart <= DATE_PROXIMITY_DAYS:
            base_score += 5  # date alignment bonus

    return min(base_score, 100.0)


def run_matching(pb_rows: list[list], pb_headers: list[str],
                 scanner_deals: list[dict]) -> tuple[list[dict], list[dict]]:
    """Match PitchBook rows against scanner deals.

    Returns:
        (matched, unmatched) — lists of enrichment result dicts.
        Each dict has: row_index, enrichment fields, match_details.
    """
    # Find column indices
    col_idx = {h: i for i, h in enumerate(pb_headers)}
    companies_col = col_idx.get("Companies", 1)
    investors_col = col_idx.get("Investors", 13)
    date_col = col_idx.get("Deal Date", 9)

    matched = []
    unmatched = []
    match_log = []

    print(f"\n  Matching {len(pb_rows)} PB rows against {len(scanner_deals)} scanner deals...")

    for row_idx, row in enumerate(pb_rows):
        pb_target = str(row[companies_col]) if row[companies_col] is not None else ""
        pb_investors = str(row[investors_col]) if investors_col < len(row) and row[investors_col] is not None else ""
        pb_acquiror = extract_acquiror_from_investors(pb_investors)
        pb_date = row[date_col] if date_col < len(row) else None

        best_score = 0.0
        best_deal = None

        for scanner_deal in scanner_deals:
            score = compute_match_score(pb_target, pb_acquiror, pb_date, scanner_deal)
            if score > best_score:
                best_score = score
                best_deal = scanner_deal

        if best_score >= HIGH_CONFIDENCE_SCORE and best_deal:
            matched.append({
                "row_index": row_idx,
                "description": best_deal.get("Description", ""),
                "opportunity": best_deal.get("Opportunity", ""),
                "verification_link": best_deal.get("Source Link", ""),
                "source_validation": best_deal.get("Source Validation", ""),
                "enrichment_source": "Scanner Match",
                "qa_status": "Pending Review",
            })
            match_log.append({
                "row_index": row_idx,
                "pb_target": pb_target,
                "pb_acquiror": pb_acquiror,
                "scanner_target": best_deal.get("Target", ""),
                "scanner_acquiror": best_deal.get("Acquiror", ""),
                "score": round(best_score, 1),
                "matched": True,
            })
        else:
            unmatched.append({
                "row_index": row_idx,
                "pb_target": pb_target,
                "pb_acquiror": pb_acquiror,
                "pb_date": str(pb_date) if pb_date else "",
                "pb_deal_type": str(row[col_idx.get("Deal Type", 10)]) if col_idx.get("Deal Type", 10) < len(row) else "",
                "pb_deal_status": str(row[col_idx.get("Deal Status", 7)]) if col_idx.get("Deal Status", 7) < len(row) else "",
                "pb_valuation": str(row[col_idx.get("Post Valuation", 8)]) if col_idx.get("Post Valuation", 8) < len(row) else "",
                "pb_investors": pb_investors,
                "pb_industry": str(row[col_idx.get("Primary PitchBook Industry Group", 15)]) if col_idx.get("Primary PitchBook Industry Group", 15) < len(row) and row[col_idx.get("Primary PitchBook Industry Group", 15)] is not None else "",
            })
            if best_score > 0:
                match_log.append({
                    "row_index": row_idx,
                    "pb_target": pb_target,
                    "pb_acquiror": pb_acquiror,
                    "best_scanner_target": best_deal.get("Target", "") if best_deal else "",
                    "best_scanner_acquiror": best_deal.get("Acquiror", "") if best_deal else "",
                    "best_score": round(best_score, 1),
                    "matched": False,
                })

    print(f"  Matched: {len(matched)}, Unmatched: {len(unmatched)}")
    return matched, unmatched, match_log


# ---------------------------------------------------------------------------
# Phase 2: AI Enrichment
# ---------------------------------------------------------------------------

DESCRIPTION_SYSTEM_PROMPT = """You are an M&A research analyst specializing in industrial software transactions.

Generate a concise 1-3 sentence deal summary with strategic rationale for the following transaction.
Focus on:
- What the target company does (products/services)
- The strategic rationale for the acquisition
- Any relevant context (market positioning, customer base)

Also extract a source URL from your web search results if available.

Return as JSON:
{{
    "description": "1-3 sentence summary",
    "source_link": "URL or null"
}}"""

OPPORTUNITY_SYSTEM_PROMPT = """You are a sales intelligence analyst for Prometheus Group (PG), which sells
maintenance planning, scheduling, and MRO (maintenance, repair, operations) software
for industrial asset management. PG's customers are heavy-industry companies —
oil & gas, mining, utilities, manufacturing, chemicals, etc.

For the following software/IT transaction, research the target company and classify
the opportunity for PG:

Target: {target}
Acquiror: {acquiror}
Deal Type: {deal_type}
Industry: {industry}
Valuation: {valuation}
Investors: {investors}

RESEARCH AND PROVIDE:

1. WHAT WAS BOUGHT/SOLD — List the target's key products, software platforms,
   intellectual property, customer base, and physical assets (offices, data centers,
   labs). Be specific: product names, number of customers, industries served,
   revenue if known. This is the most important part.

2. OPPORTUNITY CLASSIFICATION:
   - OFFENSIVE: The acquiror is expanding into industrial maintenance/asset management
     software, or the target's products overlap with PG's offerings (maintenance planning,
     scheduling, MRO, asset performance management, EAM, industrial IoT, CMMS). This
     represents competitive intelligence or partnership/acquisition opportunity for PG.
   - DEFENSIVE: The target serves PG's industrial customer base (oil & gas, mining,
     utilities, manufacturing) with complementary software, and under new ownership may
     be pulled away from PG integration/partnership. Risk to PG's ecosystem.
   - MONITOR: General market intelligence — VC rounds, deals in adjacent software markets
     (e.g., pure cybersecurity, HR tech, fintech), or transactions with no direct PG impact.

3. RECOMMENDED ACTION: 1-2 sentences for the sales team.

Return as JSON:
{{
    "classification": "OFFENSIVE|DEFENSIVE|MONITOR",
    "rationale": "1-2 sentence explanation",
    "assets_acquired": "Detailed list of products, platforms, IP, customers, and physical assets bought/sold",
    "recommended_action": "1-2 sentence action item for PG sales team"
}}"""


def enrich_single_deal(client: anthropic.Anthropic, deal_info: dict,
                       opportunity_only: bool = False) -> dict:
    """Run API calls to enrich a single unmatched deal.

    Args:
        client: Anthropic client
        deal_info: Deal metadata dict
        opportunity_only: If True, skip Call 1 (description) and only run Call 2

    Returns dict with description, opportunity, verification_link, source_validation.
    """
    target = deal_info.get("pb_target", "")
    acquiror = deal_info.get("pb_acquiror", "")
    deal_type = deal_info.get("pb_deal_type", "")
    deal_date = deal_info.get("pb_date", "")
    valuation = deal_info.get("pb_valuation", "")
    investors = deal_info.get("pb_investors", "")
    industry = deal_info.get("pb_industry", "")

    result = {
        "description": "",
        "opportunity": "",
        "verification_link": "",
        "source_validation": "Unverified",
        "enrichment_source": "AI-Generated",
        "qa_status": "Pending Review",
    }

    search_context = f"{acquiror} acquires {target}" if acquiror else f"{target} {deal_type}"
    if deal_date:
        search_context += f" {deal_date[:4]}"

    # Call 1: Description (skip if opportunity_only)
    if opportunity_only:
        result["description"] = deal_info.get("existing_description", "")
        result["verification_link"] = deal_info.get("existing_verification_link", "")
        result["source_validation"] = deal_info.get("existing_source_validation", "Unverified")
    else:
        try:
            response = call_claude_with_retry(
                client,
                model=AI_MODEL,
                max_tokens=2048,
                tools=[{
                    "type": "web_search_20250305",
                    "name": "web_search",
                    "max_uses": 3,
                }],
                system=DESCRIPTION_SYSTEM_PROMPT,
                messages=[{
                    "role": "user",
                    "content": f"Generate a deal summary for: {search_context}\nTarget: {target}\nAcquiror: {acquiror}\nDeal Type: {deal_type}\nDate: {deal_date}\nIndustry: {industry}",
                }],
            )

            data = extract_json_from_response(response)
            if isinstance(data, dict):
                result["description"] = data.get("description", "")
                link = data.get("source_link")
                if link and link != "null":
                    result["verification_link"] = link
                    result["source_validation"] = "Verified"
            elif isinstance(data, list) and data:
                result["description"] = data[0].get("description", "")

            # Fallback: extract text if JSON failed
            if not result["description"]:
                text = extract_text_from_response(response)
                if text and len(text) > 20:
                    result["description"] = text[:500]

        except Exception as e:
            if is_fatal_api_error(e):
                raise
            result["description"] = f"[Error: {str(e)[:100]}]"

        time.sleep(AI_SLEEP_BETWEEN_CALLS)

    # Call 2: Opportunity Classification
    try:
        response = call_claude_with_retry(
            client,
            model=AI_MODEL,
            max_tokens=2048,
            tools=[{
                "type": "web_search_20250305",
                "name": "web_search",
                "max_uses": 3,
            }],
            system=OPPORTUNITY_SYSTEM_PROMPT.format(
                target=target,
                acquiror=acquiror,
                deal_type=deal_type,
                industry=industry,
                valuation=valuation,
                investors=investors,
            ),
            messages=[{
                "role": "user",
                "content": f"Classify this transaction's relevance to Prometheus Group: {search_context}",
            }],
        )

        data = extract_json_from_response(response)
        if isinstance(data, dict):
            classification = data.get("classification", "MONITOR")
            rationale = data.get("rationale", "")
            assets = data.get("assets_acquired", "")
            action = data.get("recommended_action", "")
            # Format like scanner Tier 4: "CLASSIFICATION: rationale\n\nAssets: ...\n\nAction: ..."
            parts = [f"{classification}: {rationale}"]
            if assets:
                parts.append(f"\nAssets Acquired: {assets}")
            if action:
                parts.append(f"\nRecommended Action: {action}")
            result["opportunity"] = "".join(parts)
        elif isinstance(data, list) and data:
            result["opportunity"] = str(data[0])

        # Fallback
        if not result["opportunity"]:
            text = extract_text_from_response(response)
            if text:
                result["opportunity"] = text[:500]

    except Exception as e:
        if is_fatal_api_error(e):
            raise
        result["opportunity"] = f"MONITOR: [Error: {str(e)[:100]}]"

    time.sleep(AI_SLEEP_BETWEEN_CALLS)

    return result


def _worker_enrich(deal_info: dict, worker_id: int, opportunity_only: bool = False) -> tuple:
    """Worker function for parallel enrichment. Each worker gets its own client."""
    client = anthropic.Anthropic()
    row_idx = deal_info["row_index"]
    try:
        enrichment = enrich_single_deal(client, deal_info, opportunity_only=opportunity_only)
        return (row_idx, enrichment, None)
    except Exception as e:
        if is_fatal_api_error(e):
            return (row_idx, None, e)
        return (row_idx, {
            "description": f"[Error: {str(e)[:100]}]",
            "opportunity": "MONITOR: Enrichment failed",
            "verification_link": "",
            "source_validation": "Unverified",
            "enrichment_source": "AI-Generated",
            "qa_status": "Pending Review",
        }, None)


def run_ai_enrichment(client: anthropic.Anthropic, unmatched: list[dict],
                      checkpoint_path: Optional[str] = None,
                      already_enriched: Optional[dict] = None,
                      excel_callback=None,
                      num_workers: int = 1,
                      opportunity_only: bool = False) -> dict:
    """Enrich all unmatched deals via AI, optionally in parallel.

    Args:
        client: Anthropic API client (used for single-worker mode)
        unmatched: List of unmatched deal info dicts
        checkpoint_path: Path for progressive checkpointing
        already_enriched: Dict of row_index -> enrichment from resumed checkpoint
        excel_callback: Optional callable(ai_enriched_dict) to write Excel progressively
        num_workers: Number of parallel workers (default 1)
        opportunity_only: If True, only re-run opportunity classification (Call 2)

    Returns:
        Dict mapping row_index -> enrichment result dict.
    """
    results = dict(already_enriched) if already_enriched else {}
    errors = 0
    fatal = False

    # Filter out already-enriched rows
    to_process = [u for u in unmatched if u["row_index"] not in results]
    total = len(to_process)

    print(f"\n  AI Enrichment: {total} deals to process"
          f" ({len(results)} already done from checkpoint)")
    print(f"  Workers: {num_workers}")

    if num_workers <= 1:
        # Sequential mode (original behavior)
        for i, deal_info in enumerate(to_process, 1):
            row_idx = deal_info["row_index"]
            target = deal_info.get("pb_target", "?")
            acquiror = deal_info.get("pb_acquiror", "?")

            print(f"  [{i}/{total}] {acquiror} → {target}")

            try:
                enrichment = enrich_single_deal(client, deal_info, opportunity_only=opportunity_only)
                results[row_idx] = enrichment
                print(f"    {enrichment.get('opportunity', '')[:80]}")
            except Exception as e:
                if is_fatal_api_error(e):
                    print(f"\n  FATAL API ERROR: {e}")
                    print(f"  Stopping. {len(results)} deals enriched so far.")
                    fatal = True
                    break
                errors += 1
                results[row_idx] = {
                    "description": f"[Error: {str(e)[:100]}]",
                    "opportunity": "MONITOR: Enrichment failed",
                    "verification_link": "",
                    "source_validation": "Unverified",
                    "enrichment_source": "AI-Generated",
                    "qa_status": "Pending Review",
                }
                print(f"    Error: {e}")

            # Checkpoint + progressive Excel write
            if checkpoint_path and i % CHECKPOINT_EVERY_N == 0:
                save_enrichment_checkpoint(checkpoint_path, results, unmatched)
                if excel_callback:
                    excel_callback(results)
                print(f"    [Checkpoint + Excel saved: {len(results)} enriched]")
    else:
        # Parallel mode
        lock = threading.Lock()
        completed = [0]  # mutable counter

        with ThreadPoolExecutor(max_workers=num_workers) as executor:
            futures = {}
            for deal_info in to_process:
                f = executor.submit(_worker_enrich, deal_info, 0, opportunity_only)
                futures[f] = deal_info

            for future in as_completed(futures):
                deal_info = futures[future]
                row_idx, enrichment, err = future.result()

                with lock:
                    completed[0] += 1
                    count = completed[0]

                target = deal_info.get("pb_target", "?")
                acquiror = deal_info.get("pb_acquiror", "?")

                if err and is_fatal_api_error(err):
                    print(f"\n  FATAL API ERROR: {err}")
                    print(f"  Stopping. {len(results)} deals enriched so far.")
                    fatal = True
                    executor.shutdown(wait=False, cancel_futures=True)
                    break

                if enrichment:
                    with lock:
                        results[row_idx] = enrichment
                    opp = enrichment.get('opportunity', '')[:80]
                    print(f"  [{count}/{total}] {acquiror} → {target}")
                    print(f"    {opp}")
                else:
                    errors += 1
                    with lock:
                        results[row_idx] = {
                            "description": "[Error]",
                            "opportunity": "MONITOR: Enrichment failed",
                            "verification_link": "",
                            "source_validation": "Unverified",
                            "enrichment_source": "AI-Generated",
                            "qa_status": "Pending Review",
                        }
                    print(f"  [{count}/{total}] {acquiror} → {target} — ERROR")

                # Checkpoint every N completions
                if checkpoint_path and count % CHECKPOINT_EVERY_N == 0:
                    with lock:
                        save_enrichment_checkpoint(checkpoint_path, results, unmatched)
                    if excel_callback:
                        with lock:
                            excel_callback(results)
                    print(f"    [Checkpoint + Excel saved: {len(results)} enriched]")

    # Final checkpoint + Excel write
    if checkpoint_path:
        save_enrichment_checkpoint(checkpoint_path, results, unmatched)
    if excel_callback:
        excel_callback(results)

    print(f"\n  AI Enrichment complete: {len(results)} enriched, {errors} errors")
    return results


# ---------------------------------------------------------------------------
# Checkpointing
# ---------------------------------------------------------------------------

def make_checkpoint_path() -> str:
    os.makedirs(CHECKPOINT_DIR, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    return os.path.join(CHECKPOINT_DIR, f"enrich_{ts}.json")


def save_enrichment_checkpoint(path: str, enriched: dict, unmatched: list[dict]) -> None:
    """Save enrichment progress to checkpoint file."""
    # Convert int keys to strings for JSON
    serializable = {str(k): v for k, v in enriched.items()}
    data = {
        "timestamp": datetime.now().isoformat(),
        "enriched_count": len(enriched),
        "total_unmatched": len(unmatched),
        "enriched": serializable,
    }
    tmp = path + ".tmp"
    with open(tmp, "w") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)
    os.replace(tmp, path)


def load_enrichment_checkpoint(path: str) -> dict:
    """Load enrichment checkpoint, returning dict of row_index -> enrichment."""
    with open(path, "r") as f:
        data = json.load(f)

    # Convert string keys back to int
    enriched = {int(k): v for k, v in data.get("enriched", {}).items()}
    print(f"  Resumed checkpoint: {len(enriched)} deals already enriched")
    return enriched


# ---------------------------------------------------------------------------
# Phase 3: Excel Output
# ---------------------------------------------------------------------------

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
WRAP = Alignment(wrap_text=True, vertical="top")


def write_enriched_excel(input_path: str, output_path: str,
                         matched: list[dict], ai_enriched: dict) -> None:
    """Write the enriched Excel file, preserving all original PitchBook data."""
    # Load original workbook (not read_only so we can copy structure)
    wb = openpyxl.load_workbook(input_path)
    ws = wb.active

    # Build enrichment lookup: row_index -> enrichment dict
    enrichment_map = {}
    for m in matched:
        enrichment_map[m["row_index"]] = m
    for row_idx, enr in ai_enriched.items():
        if row_idx not in enrichment_map:  # matched takes priority
            enrichment_map[row_idx] = enr

    # Find the last used column (28 = AB, so new cols start at 29 = AC)
    new_col_start = 29  # Column AC (1-indexed)

    # Write new column headers in row 3
    for j, header in enumerate(NEW_HEADERS):
        cell = ws.cell(row=PB_HEADER_ROW, column=new_col_start + j, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = WRAP

    # Write enrichment data
    enriched_count = 0
    for row_idx in range(len(enrichment_map) + len(enrichment_map)):
        # row_idx is 0-based index into data rows
        # Excel row = PB_DATA_START_ROW + row_idx
        pass

    # Iterate over all data rows
    max_data_row = ws.max_row
    for excel_row in range(PB_DATA_START_ROW, max_data_row + 1):
        row_idx = excel_row - PB_DATA_START_ROW  # 0-based data index

        if row_idx in enrichment_map:
            enr = enrichment_map[row_idx]
            ws.cell(row=excel_row, column=new_col_start, value=enr.get("description", ""))
            ws.cell(row=excel_row, column=new_col_start + 1, value=enr.get("opportunity", ""))
            ws.cell(row=excel_row, column=new_col_start + 2, value=enr.get("verification_link", ""))
            ws.cell(row=excel_row, column=new_col_start + 3, value=enr.get("source_validation", ""))
            ws.cell(row=excel_row, column=new_col_start + 4, value=enr.get("enrichment_source", ""))
            ws.cell(row=excel_row, column=new_col_start + 5, value=enr.get("qa_status", "Pending Review"))
            enriched_count += 1

            # Apply wrap alignment to description and opportunity cells
            ws.cell(row=excel_row, column=new_col_start).alignment = WRAP
            ws.cell(row=excel_row, column=new_col_start + 1).alignment = WRAP

    # Auto-width for new columns
    for j, header in enumerate(NEW_HEADERS):
        col_letter = openpyxl.utils.get_column_letter(new_col_start + j)
        if header in ("Description", "Opportunity"):
            ws.column_dimensions[col_letter].width = 60
        elif header == "Verification Link":
            ws.column_dimensions[col_letter].width = 40
        else:
            ws.column_dimensions[col_letter].width = 18

    os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else ".", exist_ok=True)
    wb.save(output_path)
    print(f"\n  Saved enriched file: {output_path}")
    print(f"  Rows enriched: {enriched_count}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Enrich PitchBook Industrial Software Transactions")
    parser.add_argument("--input", required=True, help="PitchBook Excel file")
    parser.add_argument("--scanner", required=True, help="Scanner deals Excel file")
    parser.add_argument("--output", required=True, help="Output enriched Excel file")
    parser.add_argument("--skip-ai", action="store_true", help="Matching only, no API calls")
    parser.add_argument("--resume", type=str, default=None, help="Resume from checkpoint file")
    parser.add_argument("--test", type=int, default=None, help="Process first N rows only")
    parser.add_argument("--workers", type=int, default=1, help="Parallel workers (default 1)")
    parser.add_argument("--redo-opportunities", action="store_true",
                        help="Re-run only opportunity classification (keeps descriptions)")
    args = parser.parse_args()

    print("=" * 60)
    print("PitchBook Industrial Software — Deal Enrichment")
    print("=" * 60)

    # ---------------------------------------------------------------
    # Phase 1: Load data
    # ---------------------------------------------------------------
    print("\nPHASE 1: Loading data...")
    pb_rows, pb_headers = load_pitchbook(args.input)
    scanner_deals = load_scanner_deals(args.scanner)

    if args.test:
        pb_rows = pb_rows[:args.test]
        print(f"  TEST MODE: Processing first {args.test} rows only")

    # ---------------------------------------------------------------
    # Phase 1: Matching
    # ---------------------------------------------------------------
    print("\nPHASE 1: Fuzzy matching...")
    matched, unmatched, match_log = run_matching(pb_rows, pb_headers, scanner_deals)

    # Save match log
    log_path = os.path.join("data", "enrichment_match_log.json")
    os.makedirs("data", exist_ok=True)
    with open(log_path, "w") as f:
        json.dump(match_log, f, indent=2, ensure_ascii=False)
    print(f"  Match log saved: {log_path}")

    # ---------------------------------------------------------------
    # Phase 2: AI Enrichment
    # ---------------------------------------------------------------
    ai_enriched = {}

    if args.skip_ai:
        print("\nPHASE 2: SKIPPED (--skip-ai)")
    elif args.redo_opportunities:
        # Re-run only opportunity classification, keeping existing descriptions
        print("\nPHASE 2: Re-running Opportunity Classification (keeping descriptions)...")

        if not args.resume:
            print("  ERROR: --redo-opportunities requires --resume <checkpoint> to load existing descriptions")
            sys.exit(1)

        # Load existing enrichment
        existing_enriched = load_enrichment_checkpoint(args.resume)
        print(f"  Loaded {len(existing_enriched)} existing enrichments")

        # Inject existing description/link/validation into unmatched dicts
        for u in unmatched:
            ridx = u["row_index"]
            if ridx in existing_enriched:
                prev = existing_enriched[ridx]
                u["existing_description"] = prev.get("description", "")
                u["existing_verification_link"] = prev.get("verification_link", "")
                u["existing_source_validation"] = prev.get("source_validation", "Unverified")

        calls = len(unmatched)
        print(f"  {calls} deals to re-classify (~{calls} API calls)")

        client = anthropic.Anthropic()
        checkpoint_path = make_checkpoint_path()
        print(f"  Checkpoint file: {checkpoint_path}")

        def save_excel_now(current_ai_enriched):
            write_enriched_excel(args.input, args.output, matched, current_ai_enriched)

        ai_enriched = run_ai_enrichment(
            client, unmatched, checkpoint_path, already_enriched={},
            excel_callback=save_excel_now,
            num_workers=args.workers,
            opportunity_only=True
        )
    else:
        print("\nPHASE 2: AI Enrichment...")
        print(f"  {len(unmatched)} unmatched deals to enrich")
        print(f"  Estimated API calls: ~{len(unmatched) * 2}")

        client = anthropic.Anthropic()
        checkpoint_path = make_checkpoint_path()
        print(f"  Checkpoint file: {checkpoint_path}")

        # Load existing checkpoint if resuming
        already_enriched = {}
        if args.resume:
            already_enriched = load_enrichment_checkpoint(args.resume)

        # Progressive Excel writer — updates output file at each checkpoint
        def save_excel_now(current_ai_enriched):
            write_enriched_excel(args.input, args.output, matched, current_ai_enriched)

        ai_enriched = run_ai_enrichment(
            client, unmatched, checkpoint_path, already_enriched,
            excel_callback=save_excel_now,
            num_workers=args.workers
        )

    # ---------------------------------------------------------------
    # Phase 3: Write output
    # ---------------------------------------------------------------
    print("\nPHASE 3: Writing output...")
    write_enriched_excel(args.input, args.output, matched, ai_enriched)

    # ---------------------------------------------------------------
    # Summary
    # ---------------------------------------------------------------
    print("\n" + "=" * 60)
    print("ENRICHMENT SUMMARY")
    print("=" * 60)
    print(f"  Total PitchBook rows:  {len(pb_rows)}")
    print(f"  Scanner matches:       {len(matched)}")
    print(f"  AI-enriched:           {len(ai_enriched)}")
    print(f"  Not enriched:          {len(pb_rows) - len(matched) - len(ai_enriched)}")
    print(f"  Output file:           {args.output}")
    print("=" * 60)


if __name__ == "__main__":
    main()
