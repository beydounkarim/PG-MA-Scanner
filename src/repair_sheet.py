#!/usr/bin/env python3
"""
One-time repair script for the PG M&A Scanner Google Sheet.

Reads all data from the backup Excel (primary source of truth), runs fuzzy
dedup, and rewrites both the Excel and Google Sheet with date-grouped
sections and professional formatting.

Usage:
    python src/repair_sheet.py
"""

import os
import sys
import time

# Add src to path
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from dotenv import load_dotenv
load_dotenv()

from datetime import date, timedelta

import openpyxl

from sheets_output import (
    open_sheet, DEALS_HEADERS, EXCLUDED_HEADERS, UNVERIFIED_HEADERS,
    HEADER_BG, HEADER_TEXT, SECTION_BG, SECTION_TEXT,
    get_deal_date, get_date_group, get_date_group_order, get_group_header,
)
from checkpoint_manager import save_excel_backup
from dedup import fuzzy_dedupe_deals
from matcher import load_company_list, fuzzy_match

BACKUP_EXCEL = "data/PG_MA_Scanner_Deals - back-up.xlsx"

# Column mapping: Excel backup header → Google Sheets header
_DEALS_COL_MAP = {
    "Acquiror": "Acquiror",
    "Target": "Target",
    "Deal Type": "Deal Type",
    "Deal Status": "Deal Status",
    "Description": "Description",
    "Deal Value": "Deal Value ($)",
    "Date Rumored": "Date of Rumor",
    "Date Announced": "Date of Announcement",
    "Date Closed": "Date Closed",
    "Sector": "Sector",
    "Source": "Source",
    "Source Link": "Source Link",
    "PG Account": "PG Account Name",
    "Opportunity": "Potential Opportunity for PG",
    "Source Validation": "Source Validation",
}

_EXCLUDED_COL_MAP = {
    "Acquiror": "Acquiror",
    "Target": "Target",
    "Deal Value": "Deal Value ($)",
    "Sector": "Sector",
    "Description": "Description",
    "Exclusion Reason": "Exclusion Reason",
}

_UNVERIFIED_COL_MAP = {
    "Acquiror": "Acquiror",
    "Target": "Target",
    "Deal Status": "Deal Status",
    "Sector": "Sector",
    "Description": "Description",
    "Source Link": "Original URL Attempted",
    "Failure Reason": "Validation Failure Reason",
}


def _col_letter(index: int) -> str:
    """Convert 0-based column index to letter (A=0, Z=25)."""
    return chr(ord('A') + index)


def repair_sheet():
    """Main repair function: read from backup Excel, dedup, rewrite Excel + Sheet."""
    today = date.today()
    groups = get_date_group_order(today)

    print(f"Today: {today.isoformat()}")
    print(f"Date groups: {groups}\n")

    # --- Read all data from backup Excel (source of truth) ---
    print(f"Reading from backup Excel: {BACKUP_EXCEL}")
    deals_data = _read_excel_tab(BACKUP_EXCEL, "New Deals", _DEALS_COL_MAP)
    excluded_data = _read_excel_tab(BACKUP_EXCEL, "Excluded", _EXCLUDED_COL_MAP)
    unverified_data = _read_excel_tab(BACKUP_EXCEL, "Unverified", _UNVERIFIED_COL_MAP)

    print(f"\nData loaded: {len(deals_data)} deals, "
          f"{len(excluded_data)} excluded, {len(unverified_data)} unverified\n")

    # --- Fuzzy dedup deals ---
    if deals_data:
        print("=" * 60)
        print("FUZZY DEDUPLICATION")
        print("=" * 60)
        before_count = len(deals_data)
        deals_data = fuzzy_dedupe_deals(deals_data, verbose=True)
        after_count = len(deals_data)
        print(f"\n  Result: {before_count} → {after_count} deals "
              f"({before_count - after_count} removed)\n")

    # --- Fill blank Acquiror/Target from descriptions ---
    if deals_data:
        print("=" * 60)
        print("FILLING BLANK FIELDS")
        print("=" * 60)
        deals_data = _fill_blank_parties(deals_data)

    # --- Re-match PG accounts to populate PG Account Name & Clean Name ---
    if deals_data:
        print("\n" + "=" * 60)
        print("PG ACCOUNT MATCHING")
        print("=" * 60)
        deals_data = _rematch_pg_accounts(deals_data)

    # --- Write Excel (primary output) ---
    print("=" * 60)
    print("WRITING EXCEL (PRIMARY)")
    print("=" * 60)
    _regenerate_excel(deals_data, excluded_data, unverified_data)

    # --- Write Google Sheet ---
    print("\nConnecting to Google Sheets...")
    spreadsheet = open_sheet()

    print("\n" + "=" * 60)
    print("REPAIRING DEALS TAB")
    print("=" * 60)
    _repair_tab(spreadsheet, "Deals", DEALS_HEADERS, deals_data,
                today, groups, is_deals=True)

    print("\n" + "=" * 60)
    print("REPAIRING EXCLUDED TAB")
    print("=" * 60)
    _repair_tab(spreadsheet, "Excluded (Non-Strategic)", EXCLUDED_HEADERS,
                excluded_data, today, groups)

    print("\n" + "=" * 60)
    print("REPAIRING UNVERIFIED TAB")
    print("=" * 60)
    _repair_tab(spreadsheet, "Unverified", UNVERIFIED_HEADERS,
                unverified_data, today, groups)

    # --- Done ---
    print("\n" + "=" * 60)
    print("REPAIR COMPLETE")
    print("=" * 60)
    print(f"Excel: data/PG_MA_Scanner_Deals.xlsx")
    print(f"Sheet: https://docs.google.com/spreadsheets/d/{spreadsheet.id}")


def _extract_acquiror_from_desc(desc: str) -> str:
    """Extract acquiror company name from a deal description.

    Looks for patterns like:
      'CompanyName acquired/bought/completed/agreed/entered...'
      'CompanyName has acquired/entered/agreed...'
      'CompanyName provided...'
    """
    if not desc:
        return ""
    import re
    # Match: sentence-start name followed by an action verb
    # The name is everything before the verb phrase
    patterns = [
        r'^(.+?)\s+(?:acquired|completed its acquisition|completed the acquisition|'
        r'agreed to acquire|entered into|has acquired|has agreed|has entered|'
        r'has completed|is acquiring|will acquire|bought|purchases?d|'
        r'provided .+ financing|made a .+ (?:bid|offer)|announced)',
    ]
    for pat in patterns:
        m = re.match(pat, desc, re.IGNORECASE)
        if m:
            name = m.group(1).strip().rstrip(",")
            # Sanity: under 100 chars
            if len(name) < 100:
                return name
    return ""


def _fill_blank_parties(deals: list[dict]) -> list[dict]:
    """Fill blank Acquiror fields by extracting from Description."""
    filled = 0
    for deal in deals:
        acquiror = deal.get("Acquiror", "").strip()
        if not acquiror:
            desc = deal.get("Description", "")
            extracted = _extract_acquiror_from_desc(desc)
            if extracted:
                deal["Acquiror"] = extracted
                filled += 1

    blank_acq = sum(1 for d in deals if not d.get("Acquiror", "").strip())
    print(f"  Filled {filled} blank Acquiror fields from descriptions")
    print(f"  Remaining blank Acquiror: {blank_acq}")
    return deals


def _rematch_pg_accounts(deals: list[dict]) -> list[dict]:
    """Re-run fuzzy matching on all deals to populate PG Account Name & Clean Name."""
    companies = load_company_list("data/PG_Acct_List.xlsx")
    print(f"  Loaded {len(companies)} PG accounts")

    # Build direct lookup: account_name → clean_name
    acct_to_clean = {c["account_name"]: c["clean_name"] for c in companies}

    matched_fuzzy = 0
    matched_lookup = 0
    for deal in deals:
        acquiror = deal.get("Acquiror", "")
        target = deal.get("Target", "")
        existing_pg = deal.get("PG Account Name", "").strip()

        # Strategy 1: Direct lookup from existing PG Account Name
        if existing_pg and existing_pg in acct_to_clean:
            deal["PG Account Name"] = existing_pg
            deal["Clean Name"] = acct_to_clean[existing_pg]
            matched_lookup += 1
            continue

        # Strategy 2: Fuzzy match acquiror then target
        match = fuzzy_match(acquiror, companies) or fuzzy_match(target, companies)
        if match:
            deal["PG Account Name"] = match["account_name"]
            deal["Clean Name"] = match["clean_name"]
            matched_fuzzy += 1
            continue

        # Strategy 3: Keep existing PG Account, derive Clean Name from it
        if existing_pg:
            # Try partial match on the account name itself
            clean = existing_pg.split("|")[0].strip() if "|" in existing_pg else existing_pg
            deal["Clean Name"] = clean
            matched_lookup += 1

    blank_pg = sum(1 for d in deals if not d.get("PG Account Name", "").strip())
    blank_cn = sum(1 for d in deals if not d.get("Clean Name", "").strip())
    print(f"  Direct lookup: {matched_lookup}, Fuzzy matched: {matched_fuzzy}")
    print(f"  Remaining blank: {blank_pg} PG Account, {blank_cn} Clean Name\n")
    return deals


def _read_excel_tab(filepath: str, sheet_name: str,
                    col_map: dict[str, str]) -> list[dict]:
    """Read data from an Excel tab, mapping columns to Google Sheets format.

    Filters out section header rows (═══) and blank rows.
    """
    wb = openpyxl.load_workbook(filepath, read_only=True)
    if sheet_name not in wb.sheetnames:
        print(f"  Sheet '{sheet_name}' not found in Excel")
        wb.close()
        return []

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        print(f"  Sheet '{sheet_name}': empty")
        return []

    excel_headers = [str(h) if h else "" for h in rows[0]]
    records = []

    for row in rows[1:]:
        # Skip section headers
        first_cell = str(row[0]) if row[0] else ""
        if first_cell.startswith("═══"):
            continue
        # Skip empty rows
        if not any(c for c in row if c is not None and str(c).strip()):
            continue

        record = {}
        for i, excel_h in enumerate(excel_headers):
            val = row[i] if i < len(row) else None
            sheets_h = col_map.get(excel_h)
            if sheets_h:
                record[sheets_h] = str(val) if val is not None else ""
        records.append(record)

    print(f"  Sheet '{sheet_name}': read {len(records)} records")
    return records


def _read_tab_data(spreadsheet, tab_name: str) -> list[dict]:
    """Read all data records from a tab, filtering out section headers and empty rows."""
    try:
        ws = spreadsheet.worksheet(tab_name)
    except Exception:
        print(f"  Tab '{tab_name}' not found")
        return []

    all_values = ws.get_all_values()
    if len(all_values) <= 1:
        print(f"  Tab '{tab_name}': no data rows")
        return []

    header_row = all_values[0]
    records = []
    for row in all_values[1:]:
        # Skip section headers
        if row and str(row[0]).startswith("═══"):
            continue
        # Skip empty rows
        if not any(cell.strip() for cell in row):
            continue
        record = dict(zip(header_row, row))
        records.append(record)

    print(f"  Tab '{tab_name}': read {len(records)} records")
    return records


def _repair_tab(spreadsheet, tab_name: str, headers: list[str],
                records: list[dict], today: date, groups: list[str],
                is_deals: bool = False):
    """Clear a tab and rewrite with date-grouped sections and formatting."""
    try:
        ws = spreadsheet.worksheet(tab_name)
    except Exception:
        print(f"  Tab '{tab_name}' not found, creating...")
        ws = spreadsheet.add_worksheet(title=tab_name, rows=2000, cols=len(headers))

    num_cols = len(headers)
    end_col = _col_letter(num_cols - 1)

    # Group records by date
    grouped = {}
    for record in records:
        d = get_deal_date(record)
        group = get_date_group(d, today)
        grouped.setdefault(group, []).append(record)

    # Sort each group by date descending
    for group in grouped:
        grouped[group].sort(
            key=lambda r: get_deal_date(r) or date.min, reverse=True
        )

    # Build all rows: header + (section header + data) * N groups
    all_rows = [headers]
    section_header_row_nums = []  # 1-indexed row numbers

    for group in groups:
        header_text = get_group_header(group, today)
        all_rows.append([header_text] + [""] * (num_cols - 1))
        section_header_row_nums.append(len(all_rows))  # 1-indexed

        for record in grouped.get(group, []):
            row = [str(record.get(h, "")) for h in headers]
            all_rows.append(row)

    total_rows = len(all_rows)
    data_rows = total_rows - 1 - len(section_header_row_nums)

    # Clear the tab completely
    print(f"  Clearing tab...")
    ws.clear()
    time.sleep(1)

    # Remove existing banding
    try:
        sheet_props = ws.spreadsheet.fetch_sheet_metadata()
        for sheet in sheet_props.get("sheets", []):
            if sheet["properties"]["sheetId"] == ws.id:
                for br in sheet.get("bandedRanges", []):
                    spreadsheet.batch_update({"requests": [{
                        "deleteBandedRange": {
                            "bandedRangeId": br["bandedRangeId"]
                        }
                    }]})
                break
    except Exception:
        pass  # No banding to remove

    # Ensure enough rows
    if ws.row_count < total_rows + 10:
        ws.resize(rows=total_rows + 100, cols=num_cols)
        time.sleep(0.5)

    # Write all data in batches
    print(f"  Writing {total_rows} rows ({data_rows} data + "
          f"{len(section_header_row_nums)} sections)...")
    batch_size = 500
    for start in range(0, len(all_rows), batch_size):
        batch = all_rows[start:start + batch_size]
        row_start = start + 1
        row_end = row_start + len(batch) - 1
        range_str = f"A{row_start}:{end_col}{row_end}"
        ws.update(range_str, batch, value_input_option='RAW')
        time.sleep(1)

    # --- Apply formatting ---
    print(f"  Formatting header row...")
    ws.format(f"A1:{end_col}1", {
        "backgroundColor": HEADER_BG,
        "textFormat": {
            "foregroundColor": HEADER_TEXT,
            "bold": True,
            "fontSize": 10,
        },
    })
    time.sleep(0.5)

    # Format section headers
    print(f"  Formatting {len(section_header_row_nums)} section headers...")
    for row_num in section_header_row_nums:
        ws.format(f"A{row_num}:{end_col}{row_num}", {
            "backgroundColor": SECTION_BG,
            "textFormat": {
                "foregroundColor": SECTION_TEXT,
                "bold": True,
                "fontSize": 11,
            },
        })
        time.sleep(0.3)

    # Merge section header cells
    merge_requests = []
    for row_num in section_header_row_nums:
        merge_requests.append({
            "mergeCells": {
                "range": {
                    "sheetId": ws.id,
                    "startRowIndex": row_num - 1,
                    "endRowIndex": row_num,
                    "startColumnIndex": 0,
                    "endColumnIndex": num_cols,
                },
                "mergeType": "MERGE_ALL",
            }
        })

    if merge_requests:
        spreadsheet.batch_update({"requests": merge_requests})
        time.sleep(1)

    # Freeze header row
    ws.freeze(rows=1)
    time.sleep(0.5)

    # Deals tab: set column widths and hide metadata columns
    if is_deals:
        _set_deals_column_widths(spreadsheet, ws)
        _hide_metadata_columns(spreadsheet, ws)

    # Print summary
    for group in groups:
        count = len(grouped.get(group, []))
        label = get_group_header(group, today)
        print(f"    {label}: {count} deals")

    print(f"  Total: {data_rows} records written")


def _set_deals_column_widths(spreadsheet, ws):
    """Set professional column widths for the Deals tab."""
    widths = [
        (0, 1, 150),   # PG Account Name
        (1, 2, 120),   # Clean Name
        (2, 3, 200),   # Acquiror
        (3, 4, 200),   # Target
        (4, 5, 100),   # Deal Status
        (5, 6, 120),   # Sector
        (6, 7, 300),   # Description
        (7, 8, 110),   # Date of Rumor
        (8, 9, 110),   # Date of Announcement
        (9, 10, 100),  # Date Closed
        (10, 11, 100), # Deal Value
        (11, 12, 120), # Source
        (12, 13, 200), # Source Link
        (13, 14, 300), # Opportunity
        (14, 15, 120), # Source Validation
        (15, 16, 100), # Deal Type
    ]

    requests = []
    for start, end, width in widths:
        requests.append({
            "updateDimensionProperties": {
                "range": {
                    "sheetId": ws.id,
                    "dimension": "COLUMNS",
                    "startIndex": start,
                    "endIndex": end,
                },
                "properties": {"pixelSize": width},
                "fields": "pixelSize",
            }
        })

    if requests:
        spreadsheet.batch_update({"requests": requests})
        time.sleep(0.5)
        print("  Column widths set")


def _hide_metadata_columns(spreadsheet, ws):
    """Hide metadata columns Q-U (indices 16-20) on the Deals tab."""
    requests = [{
        "updateDimensionProperties": {
            "range": {
                "sheetId": ws.id,
                "dimension": "COLUMNS",
                "startIndex": 16,
                "endIndex": 21,
            },
            "properties": {"hiddenByUser": True},
            "fields": "hiddenByUser",
        }
    }]
    spreadsheet.batch_update({"requests": requests})
    time.sleep(0.5)
    print("  Metadata columns hidden")


def _regenerate_excel(deals_data: list[dict], excluded_data: list[dict],
                      unverified_data: list[dict]):
    """Regenerate Excel backup from in-memory data with date groups."""
    # Remap Google Sheets record format to pipeline dict format
    new_deals = []
    for r in deals_data:
        new_deals.append({
            "acquiror": r.get("Acquiror", ""),
            "target": r.get("Target", ""),
            "deal_type": r.get("Deal Type", ""),
            "deal_status": r.get("Deal Status", ""),
            "description": r.get("Description", ""),
            "deal_value": r.get("Deal Value ($)", ""),
            "date_rumor": r.get("Date of Rumor", ""),
            "date_announced": r.get("Date of Announcement", ""),
            "date_closed": r.get("Date Closed", ""),
            "sector": r.get("Sector", ""),
            "source": r.get("Source", ""),
            "source_link": r.get("Source Link", ""),
            "pg_account_name": r.get("PG Account Name", ""),
            "clean_name": r.get("Clean Name", ""),
            "opportunity": r.get("Potential Opportunity for PG", ""),
            "source_validation": r.get("Source Validation", ""),
        })

    excluded = []
    for r in excluded_data:
        excluded.append({
            "acquiror": r.get("Acquiror", ""),
            "target": r.get("Target", ""),
            "deal_value": r.get("Deal Value ($)", ""),
            "sector": r.get("Sector", ""),
            "description": r.get("Description", ""),
            "exclusion_reason": r.get("Exclusion Reason", ""),
        })

    unverified = []
    for r in unverified_data:
        unverified.append({
            "acquiror": r.get("Acquiror", ""),
            "target": r.get("Target", ""),
            "deal_status": r.get("Deal Status", ""),
            "sector": r.get("Sector", ""),
            "description": r.get("Description", ""),
            "source_link": r.get("Original URL Attempted", ""),
            "validation_failure_reason": r.get("Validation Failure Reason", ""),
        })

    filepath = "data/PG_MA_Scanner_Deals.xlsx"
    print(f"  Saving Excel to {filepath}...")
    save_excel_backup(new_deals, excluded, unverified, filepath)
    print(f"  Excel saved: {filepath}")


def main():
    """Main entry point."""
    print("=" * 60)
    print("PG M&A SCANNER - SHEET REPAIR")
    print("=" * 60)
    print()

    try:
        repair_sheet()
        return 0
    except Exception as e:
        print(f"\nError: {e}", file=sys.stderr)
        import traceback
        traceback.print_exc()
        return 1


if __name__ == "__main__":
    sys.exit(main())
